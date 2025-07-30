import os
import cv2
import numpy as np
import pandas as pd
import pytesseract
from datetime import datetime, timedelta
from difflib import SequenceMatcher
import time
from concurrent.futures import ThreadPoolExecutor
import multiprocessing
from openpyxl.styles import Font, Border, Side

# Version information
__version__ = "v1.0.0"

# ─── VIDEO RESOLUTION TOGGLE ────────────────────────────────────────────────────
# Set resolution: "360p", "720p", or "1080p"
RESOLUTION_SETTING = "360p"

# ─── RESOLUTION PRESETS ────────────────────────────────────────────────────────
# 360p optimized settings
SETTINGS_360P = {
    "min_area_ratio": 0.15,   # Quads must cover ≥15% of frame
    "debounce_secs": 10.0,    # Seconds between same-card logs
    "min_frames": 2,          # Persistence requirement
    "frame_skip": 4,          # Process every 4th frame for non-STOP cards
    "resize_factor": 0.4,    # Process at 40% resolution

    # Detection weights
    "w_template": 0.2,
    "w_hue": 0.1,
    "w_ocr": 0.6,
    "w_edge": 0.1,
    
    # Thresholds
    "template_strong": 0.4,
    "hue_strong": 0.6,
    "ocr_strong": 0.5,
    "edge_strong": 0.6
}

# 720p optimized settings
SETTINGS_720P = {
    "min_area_ratio": 0.20,   # Quads must cover ≥20% of frame
    "debounce_secs": 10.0,    # Seconds between same-card logs
    "min_frames": 2,          # Persistence requirement
    "frame_skip": 4,          # Process every 4th frame for non-STOP cards
    "resize_factor": 0.25,    # Process at quarter resolution
    
    # Detection weights
    "w_template": 0.2,
    "w_hue": 0.1,
    "w_ocr": 0.6,
    "w_edge": 0.1,
    
    # Thresholds
    "template_strong": 0.4,
    "hue_strong": 0.6,
    "ocr_strong": 0.5,
    "edge_strong": 0.6
}

# 1080p optimized settings (equivalent to 720p for processing efficiency)
SETTINGS_1080P = {
    "min_area_ratio": 0.20,   # Quads must cover ≥20% of frame
    "debounce_secs": 10.0,    # Seconds between same-card logs
    "min_frames": 2,          # Persistence requirement
    "frame_skip": 4,          # Process every 4th frame for non-STOP cards
    "resize_factor": 0.17,    # Process at ~17% resolution (equivalent pixel-wise to 720p at 0.25)
    
    # Detection weights
    "w_template": 0.2,
    "w_hue": 0.1,
    "w_ocr": 0.6,
    "w_edge": 0.1,
    
    # Thresholds
    "template_strong": 0.4,
    "hue_strong": 0.6,
    "ocr_strong": 0.5,
    "edge_strong": 0.6
}

# Apply selected settings
if RESOLUTION_SETTING == "360p":
    settings = SETTINGS_360P
elif RESOLUTION_SETTING == "720p":
    settings = SETTINGS_720P
elif RESOLUTION_SETTING == "1080p":
    settings = SETTINGS_1080P
else:
    raise ValueError(f"Invalid resolution setting: {RESOLUTION_SETTING}. Use '360p', '720p', or '1080p'.")

# ─── USER SETTINGS ────────────────────────────────────────────────────────────
# Change these paths:
root_folder = r"D:\record"  # Keep this the same if your videos are still here
output_excel = r"M:\Engineering\Lindsey\12. Code\Time Study Camera\TimeStudyAnalyzer\output\card_events_summary.xlsx"

# Apply resolution-specific settings
min_area_ratio = settings["min_area_ratio"]
debounce_secs  = settings["debounce_secs"]
min_frames     = settings["min_frames"]
DEBUG          = False
VERBOSE        = False  # Set to True for detailed processing info
RESIZE_FACTOR  = settings["resize_factor"]
frame_skip     = settings["frame_skip"]

# weights for combining signals (must sum to 1)
W_TEMPLATE     = settings["w_template"]
W_HUE          = settings["w_hue"]
W_OCR          = settings["w_ocr"]
W_EDGE         = settings["w_edge"]
THRESHOLD      = 0.5    # combined‐score cutoff

# any‐signal‐strong thresholds
TEMPLATE_STRONG = settings["template_strong"]
HUE_STRONG      = settings["hue_strong"]
OCR_STRONG      = settings["ocr_strong"]
EDGE_STRONG     = settings["edge_strong"]

# valid labels & their nominal hue centers & half‐ranges
VALID_WORDS = {"START", "STOP", "DOWN", "LOOK", "MEETING"}
HUE_INFO = {
    "START":   ((55+70)/2,   (70-55)/2),
    "DOWN":    ((27+33)/2,   (33-27)/2),
    "MEETING": (150.0,       15.0),   # purple: 150±15
    "LOOK":    ((92+108)/2, (108-92)/2),
    "STOP":    (0.0,         10.0),
}

# morphological kernel
kernel = np.ones((7,7), np.uint8)

# ─── LOAD TEMPLATES & EDGE TEMPLATES ─────────────────────────────────────────
script_dir      = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(script_dir, "..", ".."))
template_folder = os.path.join(project_root, "templates")
template_w, template_h = 520, 364

templates      = {}
edge_templates = {}
for lbl,(hc,hrange) in HUE_INFO.items():
    path = os.path.join(template_folder, f"{lbl}.png")
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Missing template: {path}")
    gray = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
    gray = cv2.resize(gray, (template_w, template_h), interpolation=cv2.INTER_AREA)
    templates[lbl]      = gray
    edge_templates[lbl] = cv2.Canny(gray, 50, 150)

# ─── DEBUG WINDOWS ────────────────────────────────────────────────────────────
if DEBUG:
    cv2.namedWindow("Debug", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("Debug", 800, 600)
    for lbl in HUE_INFO:
        cv2.namedWindow(f"mask_{lbl}", cv2.WINDOW_NORMAL)
        cv2.resizeWindow(f"mask_{lbl}", 300, 300)
    cv2.namedWindow("Edges", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("Edges", 300, 300)

# ─── STATE ─────────────────────────────────────────────────────────────────────
detections    = []
last_detected = {lbl: datetime.min for lbl in HUE_INFO}
consec_counts = {lbl: 0 for lbl in HUE_INFO}
currently_detected = {lbl: False for lbl in HUE_INFO}
stop_active = False  # Tracks if STOP is currently detected
last_global_detected = datetime.min
global_debounce_secs = 10.0  # or whatever you want
stop_consec = 0  # Already declared but not used

# Add this function to process each clip independently
def process_clip(clip_info):
    """Process a single video clip and return its detections"""
    date_f, hr, clip_f, clip_path, clip_start = clip_info
    
    # Initialize local state for this clip
    local_detections = []
    last_detected = {lbl: datetime.min for lbl in HUE_INFO}
    consec_counts = {lbl: 0 for lbl in HUE_INFO}
    stop_active = False
    last_global_detected = datetime.min
    stop_consec = 0
    
    print(f"▶️ Processing {date_f}/{hr}/{clip_f}")
    cap = cv2.VideoCapture(clip_path)
    if not cap.isOpened():
        print(f"⚠️ Can't open {clip_path}")
        return []
        
    clip_start_time = time.time()
    frames_in_clip = 0
        
    w_frame = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h_frame = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    frame_area = w_frame * h_frame
    
    # Reset frame count for each clip
    frame_count = 0
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        frames_in_clip += 1

        # --- Get frame timestamp ---
        ms = cap.get(cv2.CAP_PROP_POS_MSEC)
        vid_secs = ms / 1000.0
        real_ts = clip_start + timedelta(seconds=vid_secs)
        
        # OPTIMIZATION: Resize frame for faster processing
        if RESIZE_FACTOR != 1.0:
            frame = cv2.resize(frame, (0, 0), fx=RESIZE_FACTOR, fy=RESIZE_FACTOR)
            w_frame_resized = int(w_frame * RESIZE_FACTOR)
            h_frame_resized = int(h_frame * RESIZE_FACTOR) 
            frame_area = w_frame_resized * h_frame_resized
        
        # --- IMMEDIATE STOP DETECTION (every frame) ---
        hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        stop_mask = bitwise_or(
            inRange(hsv, COLOR_RANGES["STOP"][0], COLOR_RANGES["STOP"][1]),
            inRange(hsv, COLOR_RANGES["STOP"][2], COLOR_RANGES["STOP"][3])
        )
        
        # Find STOP quads
        stop_cnts, _ = findContours(stop_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        stop_valid_quads = []
        
        # OPTIMIZATION: Cache contourArea results
        stop_areas = {}
        for c in stop_cnts:
            area = contourArea(c)
            stop_areas[id(c)] = area
            if area < 0.20 * frame_area:  # 20% threshold
                continue
                
            peri = arcLength(c, True)
            approx = approxPoly(c, peri)
            if len(approx) == 4:
                # Add solidity check
                x, y, w, h = boundingRect(c)
                if area/(w*h) < 0.70:  # Ensure solid rectangle - using cached area
                    continue
                stop_valid_quads.append(c)

        stop_valid_cnt = max(stop_valid_quads, key=lambda c: stop_areas[id(c)]) if stop_valid_quads else None

        if stop_valid_cnt is not None:
            stop_consec += 1
            if stop_consec >= 2:  # Reduced to 2 consecutive frames
                if not stop_active:
                    prev = last_detected["STOP"]
                    if (real_ts - prev).total_seconds() > debounce_secs:
                        local_detections.append({
                            "Label": "STOP", 
                            "Video Time": f"{vid_secs:.2f}",
                            "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                        })
                        last_detected["STOP"] = real_ts
                        last_global_detected = real_ts  # Update global debounce
                        for k in consec_counts: consec_counts[k]=0
                        
                        # Debug info
                        if DEBUG:
                            print(f"STOP detected at {real_ts.strftime('%H:%M:%S')}")
                stop_active = True
                
                # Display frame with annotation
                if DEBUG:
                    x, y, w, h = cv2.boundingRect(stop_valid_cnt)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
                    cv2.putText(frame, "STOP", (x, y-10),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
                    cv2.imshow("Debug", frame)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        cap.release()
                        cv2.destroyAllWindows()
                        exit(0)
                        
                continue  # Skip other card processing
        else:
            stop_consec = 0
            stop_active = False

        # --- FRAME SKIPPING FOR OTHER CARDS ---
        frame_count += 1
        if frame_count % frame_skip != 0:
            continue  # Skip non-STOP processing for this frame
        
        # --- GLOBAL DEBOUNCE CHECK ---
        if (real_ts - last_global_detected).total_seconds() < global_debounce_secs:
            continue  # Skip if we recently detected any card
        
        # --- PROCESS OTHER CARDS in priority order ---
        # Process cards in strict priority order
        for lbl in ["MEETING", "START", "DOWN", "LOOK"]:  # Priority order
            # Skip STOP as it's already handled
            hc, hrange = HUE_INFO[lbl]
            
            # thresholds - can be customized per label
            area_thresh = 0.10 if lbl == "START" else min_area_ratio
            frame_thresh = 1 if lbl == "START" else min_frames
            
            # Build mask for this label
            if lbl == "MEETING":
                # MEETING mask
                hsv_mask = inRange(hsv, COLOR_RANGES["MEETING"][0], COLOR_RANGES["MEETING"][1])
                raw_mask = bitwise_and(hsv_mask, bitwise_not(stop_mask))
                v = hsv[:,:,2]
                _, v_mask = threshold(v, 80, 255, cv2.THRESH_BINARY)
                raw_mask = bitwise_and(raw_mask, v_mask)
            else:
                # Other labels masks
                lo, hi = COLOR_RANGES[lbl]
                raw_mask = inRange(hsv, lo, hi)
            
            # Single morphology operation
            mask = morph_close(raw_mask)
            if DEBUG:
                cv2.imshow(f"mask_{lbl}", mask)

            # Find valid quads
            cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            valid_quads = []
            
            # OPTIMIZATION: Cache contourArea results
            contour_areas = {}
            for c in cnts:
                area = cv2.contourArea(c)
                contour_areas[id(c)] = area
                if area < area_thresh * frame_area:
                    continue
                peri = cv2.arcLength(c, True)
                approx = cv2.approxPolyDP(c, 0.04*peri, True)
                if len(approx) == 4:
                    valid_quads.append(c)
    
            valid_cnt = max(valid_quads, key=lambda c: contour_areas[id(c)]) if valid_quads else None
    
            # MEETING-specific checks
            if lbl == "MEETING" and valid_cnt is not None:
                area = contour_areas[id(valid_cnt)]
                if area > 0.6 * frame_area:
                    valid_cnt = None  # Likely a STOP
    
            # Update consecutive counter
            consec_counts[lbl] = consec_counts[lbl] + 1 if valid_cnt is not None else 0
    
            # If not enough consecutive frames, skip to next label
            if consec_counts[lbl] < frame_thresh or valid_cnt is None:
                continue
            
            # We have a valid candidate that meets persistence requirements
            x,y,w_bb,h_bb = cv2.boundingRect(valid_cnt)
            roi = frame[y:y+h_bb, x:x+w_bb]
            
            # OCR preparation
            gray_roi = cvtColor(roi, cv2.COLOR_BGR2GRAY)
            _, roi_bin = threshold(gray_roi, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
            roi_rs = resize(roi_bin, (template_w, template_h), interpolation=cv2.INTER_AREA)
            
            # OCR
            ocr_raw = pytesseract.image_to_string(
                roi_rs,
                config="--psm 8 --oem 3 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            )
            ocr_text = "".join(ch for ch in ocr_raw if ch.isalpha()).upper()
            ocr_ratio = hamming_ratio(ocr_text, lbl) if ocr_text in VALID_WORDS else 0.0
            
            # Strong OCR detection - accept immediately
            if ocr_text in VALID_WORDS and ocr_ratio >= OCR_STRONG:
                # Special case for MEETING/STOP based on letter count
                if ocr_text == "STOP" or (lbl == "MEETING" and len(ocr_text) <= 4):
                    detected_label = "STOP"
                else:
                    detected_label = lbl
                    
                prev = last_detected[detected_label]
                if (real_ts - prev).total_seconds() > debounce_secs:
                    local_detections.append({
                        "Label": detected_label,
                        "Video Time": f"{vid_secs:.2f}",  # Format as seconds with 2 decimal places
                        "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                    })
                    last_detected[detected_label] = real_ts
                    last_global_detected = real_ts
                    for k in consec_counts: consec_counts[k]=0
                break  # Short-circuit - skip checking other labels
            
            # Fallback scoring if OCR wasn't strong enough
            roi_edges = cv2.Canny(roi_rs, 50, 150)
            gray_corr = cv2.matchTemplate(roi_rs, templates[lbl], cv2.TM_CCOEFF_NORMED).max()
            edge_corr = cv2.matchTemplate(roi_edges, edge_templates[lbl], cv2.TM_CCOEFF_NORMED).max()
            avg_hue = float(np.mean(hsv[y:y+h_bb, x:x+w_bb,0]))
            diff = abs(avg_hue - hc)
            if lbl=="STOP": diff = min(diff,180-diff)
            hue_score = max(0.0, 1.0 - diff/hrange)
            
            final_score = (
                W_TEMPLATE*gray_corr +
                W_EDGE*edge_corr +
                W_HUE*hue_score +
                W_OCR*ocr_ratio
            )
            
            # Accept if any scoring method is strong enough
            if (gray_corr >= TEMPLATE_STRONG or 
                edge_corr >= EDGE_STRONG or 
                final_score >= THRESHOLD or 
                hue_score >= HUE_STRONG):
                
                prev = last_detected[lbl]
                if (real_ts - prev).total_seconds() > debounce_secs:
                    local_detections.append({
                        "Label": lbl,
                        "Video Time": f"{vid_secs:.2f}",  # Format as seconds with 2 decimal places
                        "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                    })
                    last_detected[lbl] = real_ts
                    last_global_detected = real_ts
                    for k in consec_counts: consec_counts[k]=0
                break  # Short-circuit - skip checking other labels
        
        # Debug display
        if DEBUG:
            cv2.imshow("Debug", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                cap.release()
                cv2.destroyAllWindows()
                exit(0)

    cap.release()
    
    # Show clip stats only if verbose
    clip_elapsed = time.time() - clip_start_time
    fps = frames_in_clip / clip_elapsed if clip_elapsed > 0 else 0
    if VERBOSE:  # Only print if verbose mode is enabled
        print(f"    ✓ Processed {frames_in_clip} frames in {clip_elapsed:.2f}s ({fps:.1f} fps)")
    
    return local_detections

def auto_adjust_xlsx_columns(writer, sheet_name, df):
    """
    Auto-adjust columns width in an Excel worksheet to fit the content.
    Also:
    - Removes all cell borders
    - Centers the table on the page for printing
    - Keeps headers bold
    - Applies proper formatting to date and time columns
    """
    worksheet = writer.sheets[sheet_name]
    
    # Auto-adjust columns based on DataFrame content width
    for idx, col in enumerate(df.columns):
        # Find the maximum length in the column
        max_len = max(
            # Column name length
            len(str(col)),
            # Maximum content length
            df[col].astype(str).str.len().max() if not df.empty else 0
        )
        
        # Add some padding
        adjusted_width = max_len + 2
        
        # Limit width to a reasonable maximum
        if adjusted_width > 50:
            adjusted_width = 50
        
        # Set the column width
        col_letter = chr(65 + idx) if idx < 26 else chr(64 + idx//26) + chr(65 + idx%26)
        worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Apply proper formatting to Time column
        if col == "Time":
            for row_idx, _ in enumerate(df.index, start=2):
                cell = f"{col_letter}{row_idx}"
                worksheet[cell].number_format = "hh:mm:ss"
                
        # Apply proper formatting to Date column
        elif col == "Date":
            for row_idx, _ in enumerate(df.index, start=2):
                cell = f"{col_letter}{row_idx}"
                worksheet[cell].number_format = "mm/dd/yyyy"

    # Remove borders from all cells
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    
    # Apply formatting to all cells
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)  # Fixed: column not col
            cell.border = no_border
            
            # Keep headers bold
            if row == 1:
                cell.font = Font(bold=True)
    
    # Center the table for printing
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = True
    worksheet.page_setup.fitToHeight = True
    worksheet.print_options.horizontalCentered = True
    worksheet.print_options.verticalCentered = True

def add_time_borders(worksheet, df, columns):
    """
    Add borders to separate detections by hour and day.
    - Regular bottom border when hour changes
    - Thick bottom border when day changes
    """
    # Skip if empty dataframe
    if df.empty:
        return
        
    # Create border styles
    thin_border = Border(
        bottom=Side(style='thin')
    )
    thick_border = Border(
        bottom=Side(style='thick')
    )
    
    # Sort by time to ensure chronological order
    df = df.sort_values("Real Time")
    
    # Get the total number of columns with data
    max_col = len(columns)
    
    # Start from row 2 (first data row after header)
    prev_hour = df.iloc[0]["Real Time"].hour
    prev_day = df.iloc[0]["Real Time"].day
    
    # Iterate through rows (starting from index 1, row 3 in Excel)
    for i in range(1, len(df)):
        current_hour = df.iloc[i]["Real Time"].hour
        current_day = df.iloc[i]["Real Time"].day
        excel_row = i + 2  # +2 because Excel is 1-indexed and we have a header row
        
        # Check for day change (stronger visual separation)
        if current_day != prev_day:
            # Add thick border to previous row
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=excel_row-1, column=col)
                cell.border = thick_border
            prev_hour = current_hour
            prev_day = current_day
            
        # Check for hour change
        elif current_hour != prev_hour:
            # Add thin border to previous row
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=excel_row-1, column=col)
                cell.border = thin_border
            prev_hour = current_hour

# ─── PRECOMPUTE FUNCTIONS AND COLOR RANGES ─────────────────────────────────────
# Function references to avoid repeated lookups
inRange = cv2.inRange
bitwise_or = cv2.bitwise_or
bitwise_and = cv2.bitwise_and
bitwise_not = cv2.bitwise_not
findContours = cv2.findContours
contourArea = cv2.contourArea
arcLength = cv2.arcLength
boundingRect = cv2.boundingRect
threshold = cv2.threshold
resize = cv2.resize
Canny = cv2.Canny
matchTemplate = cv2.matchTemplate
cvtColor = cv2.cvtColor

# Create optimized operation functions
morph_close = lambda m: cv2.morphologyEx(m, cv2.MORPH_CLOSE, kernel)
approxPoly = lambda c, peri: cv2.approxPolyDP(c, 0.04*peri, True)

# Precompute color ranges for faster lookup
COLOR_RANGES = {
    "STOP": [(0,120,120), (8,255,255), (172,120,120), (180,255,255)],  # Two ranges for red
    "START": [(45,80,80), (80,255,255)],
    "DOWN": [(20,140,140), (40,255,255)],  # Wider range for better detection
    "LOOK": [(88,150,150), (112,255,255)],
    "MEETING": [(135,80,80), (165,255,255)]
}

def hamming_ratio(a, b):
    """Fast string similarity for card text comparison"""
    # Handle length differences (important for OCR results)
    len_a, len_b = len(a), len(b)
    if len_a == 0 or len_b == 0:
        return 0.0
    
    # Apply length penalty (reduces score for different length strings)
    length_factor = min(len_a, len_b) / max(len_a, len_b)
    
    # Count matching characters at same positions
    matches = sum(1 for x, y in zip(a, b) if x == y)
    
    # Normalize by max length and apply length penalty
    return (matches / max(len_a, len_b)) * length_factor

def main():
    # Start timer for overall processing
    start_time = time.time()
    
    # Collect all clips to process
    clips_to_process = []
    
    for date_f in sorted(os.listdir(root_folder)):
        date_path = os.path.join(root_folder, date_f)
        if not os.path.isdir(date_path) or not date_f.isdigit():
            continue

        for hr in sorted(os.listdir(date_path)):
            hr_path = os.path.join(date_path, hr)
            if not os.path.isdir(hr_path) or not hr.isdigit():
                continue

            for clip_f in sorted(os.listdir(hr_path)):
                if not clip_f.endswith(".mp4") or not clip_f[:-4].isdigit():
                    continue
                    
                clip_date = datetime.strptime(date_f, "%Y%m%d")
                clip_start = clip_date + timedelta(
                    hours=int(hr), minutes=int(clip_f[:-4])
                )
                clip_path = os.path.join(hr_path, clip_f)
                
                clips_to_process.append((date_f, hr, clip_f, clip_path, clip_start))
    
    # Process clips in parallel
    all_detections = []
    total_videos_processed = len(clips_to_process)
    
    # OPTIMIZATION: More intelligent worker allocation
    total_cpus = multiprocessing.cpu_count()
    # Use max 60% of CPUs
    max_workers_conservative = max(1, int(total_cpus * 0.60))
    num_workers = min(max_workers_conservative, len(clips_to_process))
    print(f"\n🖥️ Processing {total_videos_processed} videos using {num_workers}/{total_cpus} CPU cores (threaded, conservative)")
    
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        results = list(executor.map(process_clip, clips_to_process))
        
    # Combine all results
    for clip_detections in results:
        all_detections.extend(clip_detections)
    
    # Complete processing time
    total_elapsed = time.time() - start_time
    
    print("\n==== Processing Complete ====")
    print(f"Total time: {total_elapsed:.2f} seconds")
    print(f"Videos processed: {total_videos_processed}")
    print(f"Detections found: {len(all_detections)}")
    
    # ─── EXPORT TO EXCEL ─────────────────────────────────────────────────────────
    if all_detections:
        # Process raw data
        df = pd.DataFrame(all_detections)
        df["Real Time"] = pd.to_datetime(df["Real Time"])
        
        # Change this section to use the proper date format
        df["Date"] = df["Real Time"].dt.strftime("%m/%d/%Y")  # M/D/YYYY format (7/25/2025)
        df["Time"] = df["Real Time"].dt.strftime("%H:%M:%S")  # 24-hour format (14:12:00)
        
        # Sort by time to ensure chronological order
        df = df.sort_values("Real Time")
        
        # Extract LOOK events with separated date/time
        look_events = df[df["Label"] == "LOOK"].copy()
        
        # Track state transitions for accurate timing - SINGLE STATE TRACKING SYSTEM
        current_state = None
        state_start_time = None
        state_durations = {
            "START": 0,
            "MEETING": 0, 
            "DOWN": 0
        }

        # Process each event chronologically
        for i, event in df.iterrows():
            label = event["Label"]
            timestamp = event["Real Time"]
            
            # LOOK cards don't affect timing
            if label == "LOOK":
                continue
            
            # Calculate duration for the current state ALWAYS
            if current_state is not None and state_start_time is not None:
                duration = (timestamp - state_start_time).total_seconds()
                if current_state in state_durations:  # Only add to tracked states
                    state_durations[current_state] += duration
                    if VERBOSE:  # Only print if verbose mode is enabled
                        print(f"Adding {duration:.1f}s to {current_state}")
            
            # Update state based on the card
            if label == "STOP":
                current_state = None  # STOP ends timing but doesn't start a new state
                state_start_time = None
            else:  # START, MEETING, or DOWN - always update start time
                current_state = label
                state_start_time = timestamp

        # Add this section to handle the final state
        if current_state is not None and state_start_time is not None:
            # Get the last timestamp of ANY card detection - that's our true end
            last_detected_time = df["Real Time"].max()
            
            # Only count time until the last detection
            final_duration = (last_detected_time - state_start_time).total_seconds()
            
            # If positive, add it to the appropriate state's duration
            if final_duration > 0 and current_state in state_durations:
                state_durations[current_state] += final_duration
                if VERBOSE:  # Only print if verbose mode is enabled
                    print(f"Added final {current_state} time: {timedelta(seconds=final_duration)}")

        # Create summary statistics
        stats = []
        stats.append({"Metric": "Total START Time", "Value": str(timedelta(seconds=state_durations["START"]))})
        stats.append({"Metric": "Total DOWN Time", "Value": str(timedelta(seconds=state_durations["DOWN"]))})
        stats.append({"Metric": "Total MEETING Time", "Value": str(timedelta(seconds=state_durations["MEETING"]))})
        stats_df = pd.DataFrame(stats)
        
        # TAB 2: Extract LOOK events with separated date/time
        look_events = df[df["Label"] == "LOOK"].copy()
        
        # Prepare Excel writer
        os.makedirs(os.path.dirname(output_excel), exist_ok=True)
        if os.path.exists(output_excel):
            os.remove(output_excel)
            
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            # Tab 1: Summary statistics
            stats_df.to_excel(writer, sheet_name="Summary Statistics", index=False)
            auto_adjust_xlsx_columns(writer, "Summary Statistics", stats_df)
            
            # Tab 2: LOOK events
            if not look_events.empty:
                # Add video file links for LOOK events
                look_events["Video File"] = look_events.apply(
                    lambda row: f'=HYPERLINK("D:\\record\\{row["Real Time"].strftime("%Y%m%d")}\\{row["Real Time"].strftime("%H")}\\{row["Real Time"].strftime("%M")}.mp4", "Open Video")',
                    axis=1
                )
                
                # Add the Video File column to the output
                look_events.to_excel(writer, sheet_name="LOOK Events", 
                               index=False, columns=["Date", "Time", "Video Time", "Video File"])
                auto_adjust_xlsx_columns(writer, "LOOK Events", 
                               look_events[["Date", "Time", "Video Time", "Video File"]])
            else:
                # Make sure empty dataframe has all columns
                empty_df = pd.DataFrame(columns=["Date", "Time", "Video Time", "Video File"])
                empty_df.to_excel(writer, sheet_name="LOOK Events", index=False)
                auto_adjust_xlsx_columns(writer, "LOOK Events", empty_df)
            
            # Tab 3: Raw data (all detections with split date/time)
            # Create video file links
            df["Video File"] = df.apply(
                lambda row: f'=HYPERLINK("D:\\record\\{row["Real Time"].strftime("%Y%m%d")}\\{row["Real Time"].strftime("%H")}\\{row["Real Time"].strftime("%M")}.mp4", "Open Video")',
                axis=1
            )

            # Add the column to the output
            columns = ["Label", "Date", "Time", "Video Time", "Video File"]
            df.to_excel(writer, sheet_name="Raw Detections", index=False, columns=columns)
            auto_adjust_xlsx_columns(writer, "Raw Detections", df[columns])
            add_time_borders(writer.sheets["Raw Detections"], df, columns)

        print("\n✅ Enhanced report saved to", output_excel)
    else:
        print("\n⚠️ No detections found.")
        
    # OPTIMIZATION: Release memory after processing
    cv2.destroyAllWindows()
# Memory cleanup is handled automatically by Python's garbage collector

if __name__ == "__main__":
    main()

