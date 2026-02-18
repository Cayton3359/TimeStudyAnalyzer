import cv2
import numpy as np
import pytesseract
import time
import os
import pandas as pd
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
import multiprocessing
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# Import base analyzer and reuse its assets/config/export
try:
    from . import analyze_cards as base
except ImportError:
    import sys
    import importlib.util
    import os

    # If the base analyzer is already loaded (e.g. by the GUI), reuse it so
    # updated settings (paths, resolution config) carry through.
    _existing = sys.modules.get("analyze_cards")
    if _existing is not None:
        base = _existing
    else:
        _base_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "analyze_cards.py")
        _spec = importlib.util.spec_from_file_location("analyze_cards", _base_path)
        if _spec is None or _spec.loader is None:
            raise ImportError(f"Failed to load analyzer module from {_base_path}")
        base = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(base)

__version__ = base.__version__ + "-fast"

# Fast-pass tunables
NEAR_HIT_FOCUS_SECS = 1.0      # After a near hit, process every frame temporarily (REDUCED to 1 sec)
POST_CONFIRM_FOCUS_SECS = 1.0  # After a confirmed card, process every frame briefly (REDUCED to 1 sec)
NEAR_HIT_DEBOUNCE_SECS = 2.0   # Minimum spacing between near-hit logs
BASELINE_SKIP_STOP = 8         # Process 1 of every N frames after STOP
BASELINE_SKIP_OTHER = 8        # Process 1 of every N frames after other cards
NEAR_HIT_SUPPRESSION_WINDOW_SECS = 3.0  # Suppress near hits if confirmed within this window
GLARE_CONFIRMATION_DEBOUNCE_SECS = 4.0  # Remove glare entries if confirmation within this window
OCR_FUZZY = 0.55               # Accept fuzzy OCR when hue is strong
LOOSE_HUE_FOCUS = 0.30         # Loose hue trigger for focus mode

# Detection strength thresholds (lowered for higher sensitivity)
OCR_STRONG = 0.35              # Confidence threshold for exact OCR match (lowered)
HUE_STRONG = 0.35              # Hue strength threshold (lowered)
TEMPLATE_STRONG = 0.20         # Template matching sensitivity (lowered)

# Confirmation requirements (lowered for higher sensitivity)
MIN_AREA_RATIO = 0.08          # Quads must cover at least this % of frame (lowered)
MIN_FRAMES = 1                 # Only 1 consecutive frame required for confirmation (more sensitive)
CONFIRMATION_DEBOUNCE_SECS = 5.0  # Seconds between same-card confirmations

# Video processing compression factors (for each resolution when compress is enabled)
RESIZE_FACTOR_360P = 0.4       # 360p resize factor (1.0 = no compression)
RESIZE_FACTOR_720P = 0.25      # 720p resize factor (quarter resolution)
RESIZE_FACTOR_1080P = 0.17     # 1080p resize factor (equivalent pixel-wise to 720p at 0.25)

def _in_range(hsv, lo, hi):
    return cv2.inRange(hsv, np.array(lo, dtype=np.uint8), np.array(hi, dtype=np.uint8))

def _lev_ratio(a: str, b: str) -> float:
    """Normalized Levenshtein similarity ratio between strings a and b (0..1)."""
    a = a or ""
    b = b or ""
    la, lb = len(a), len(b)
    if la == 0 and lb == 0:
        return 1.0
    # Edit distance DP
    dp = [[0]*(lb+1) for _ in range(la+1)]
    for i in range(la+1):
        dp[i][0] = i
    for j in range(lb+1):
        dp[0][j] = j
    for i in range(1, la+1):
        for j in range(1, lb+1):
            cost = 0 if a[i-1] == b[j-1] else 1
            dp[i][j] = min(
                dp[i-1][j] + 1,       # deletion
                dp[i][j-1] + 1,       # insertion
                dp[i-1][j-1] + cost   # substitution
            )
    dist = dp[la][lb]
    return max(0.0, 1.0 - (dist / max(la, lb)))

def _ocr_signature_bonus(text: str, label: str) -> float:
    """Heuristic bonus based on distinctive letter patterns per label.
    - LOOK: contains 'OO' adjacent
    - MEETING: contains 'M' and preferably two 'E's
    - DOWN: contains 'D'
    - STOP: contains 'P'
    - START: contains 'AR'
    Returns a small bonus (0..0.3) to boost fuzzy OCR confidence.
    """
    if not text:
        return 0.0
    t = text.upper()
    bonus = 0.0
    if label == "LOOK":
        if "OO" in t:
            bonus += 0.20
        if "LO" in t:
            bonus += 0.10
        if "OK" in t:
            bonus += 0.10
        if "OOK" in t:
            bonus += 0.10
    if label == "MEETING":
        if "M" in t:
            bonus += 0.15
        if t.count("E") >= 2:
            bonus += 0.15
        if "ING" in t:
            bonus += 0.15
    if label == "DOWN":
        if "D" in t:
            bonus += 0.15
        if "W" in t:
            bonus += 0.10
        if "DO" in t:
            bonus += 0.10
        if "OW" in t:
            bonus += 0.10
        if "WN" in t:
            bonus += 0.10
    if label == "STOP":
        if "P" in t:
            bonus += 0.15
        if "ST" in t:
            bonus += 0.10
        if "OP" in t:
            bonus += 0.10
    if label == "START":
        if "AR" in t:
            bonus += 0.15
        if "RT" in t:
            bonus += 0.15
        if "ST" in t:
            bonus += 0.10
        if "TA" in t:
            bonus += 0.10
    return min(bonus, 0.30)

def _ocr_signature_penalty(text: str, label: str) -> float:
    """Light penalties when strongly expected letters are missing (helps tie-break).
    Caps at 0.15 to avoid overpowering positives.
    """
    if not text:
        return 0.0
    t = text.upper()
    pen = 0.0
    if label == "LOOK" and "K" not in t:
        pen += 0.05
    if label == "STOP" and "T" not in t:
        pen += 0.05
    if label == "MEETING":
        vowels = sum(ch in "AEIOU" for ch in t)
        if vowels < 2:
            pen += 0.05
    return min(pen, 0.15)

def _calculate_glare_ratio(hsv_roi: np.ndarray) -> float:
    """Calculate mean brightness (V) in an ROI (0.0 to 1.0)."""
    if hsv_roi.size == 0:
        return 0.0
    v = hsv_roi[:, :, 2]
    return float(np.mean(v)) / 255.0


def process_clip(clip_info):
    """Fast-pass version of process_clip with dynamic skip and near-hit focus.
    Reuses base module constants, templates, and export format.
    """
    date_f, hr, clip_f, clip_path, clip_start = clip_info

    local_detections = []
    local_near_hits = []
    last_detected = {lbl: datetime.min for lbl in base.HUE_INFO}
    consec_counts = {lbl: 0 for lbl in base.HUE_INFO}
    stop_active = False
    last_global_detected = datetime.min
    stop_consec = 0
    
    # Track glare delta (change) to detect sudden spikes, not persistent bright areas
    previous_glare_ratio = {lbl: 0.0 for lbl in base.HUE_INFO}
    previous_glare_ratio["__frame__"] = 0.0
    glare_delta_threshold = 0.15  # Mean V delta (0..1) required to trigger glare
    frame_glare_initialized = False
    
    # Optimization: Index near-hits by label for O(1) lookup instead of O(n) iteration
    recent_near_hits_by_label = {lbl: [] for lbl in base.HUE_INFO}
    recent_near_hits_by_label["__frame__"] = []
    last_glare_entry_index = None
    last_glare_ts = None
    last_glare_entry_index = None
    last_glare_ts = None
    last_confirmation_ts = datetime.min  # Track last confirmation to suppress glare entries

    def _has_recent_near_hit(label, ts):
        """Fast lookup of recent near-hits by label (O(1) dict access + cleanup)."""
        cutoff = ts - timedelta(seconds=NEAR_HIT_SUPPRESSION_WINDOW_SECS)
        if label in recent_near_hits_by_label:
            recent_near_hits_by_label[label] = [nh_ts for nh_ts in recent_near_hits_by_label[label] if nh_ts >= cutoff]
            return len(recent_near_hits_by_label[label]) > 0
        return False

    def _maybe_log_detect_near_hit(label, ts, vid_secs, best_guess, reason):
        if _has_recent_near_hit(label, ts):
            return
        local_near_hits.append({
            "Label": "NEED CONFIRMATION",
            "Closest Label": label,
            "Video Time": f"{vid_secs:.2f}",
            "Real Time": ts.strftime("%Y-%m-%d %H:%M:%S"),
            "Best Guess": best_guess or label,
            "Reason": f"near_hit_on_detect {reason}"
        })

    def _record_glare_entry(entry_index, ts):
        nonlocal last_glare_entry_index, last_glare_ts
        last_glare_entry_index = entry_index
        last_glare_ts = ts

    def _override_recent_glare(label, ts):
        nonlocal last_glare_entry_index, last_glare_ts
        if last_glare_entry_index is None or last_glare_ts is None:
            return
        if (ts - last_glare_ts).total_seconds() > NEAR_HIT_SUPPRESSION_WINDOW_SECS:
            return
        if 0 <= last_glare_entry_index < len(local_near_hits):
            entry = local_near_hits[last_glare_entry_index]
            if entry.get("Label") == "GLARE":
                entry["Label"] = "NEED CONFIRMATION"
                entry["Closest Label"] = label
                entry["Reason"] = "glare_overridden_by_near_hit"
        last_glare_entry_index = None
        last_glare_ts = None

    def _remove_recent_glare(ts):
        nonlocal last_glare_entry_index, last_glare_ts, last_confirmation_ts
        last_confirmation_ts = ts  # Track this confirmation
        if last_glare_entry_index is None or last_glare_ts is None:
            return
        if (ts - last_glare_ts).total_seconds() > GLARE_CONFIRMATION_DEBOUNCE_SECS:
            return
        if 0 <= last_glare_entry_index < len(local_near_hits):
            entry = local_near_hits[last_glare_entry_index]
            if entry.get("Label") == "GLARE":
                local_near_hits.pop(last_glare_entry_index)
        last_glare_entry_index = None
        last_glare_ts = None

    # Focus & debounce state
    focus_active_until = datetime.min
    last_near_hit_ts_by_label = {lbl: datetime.min for lbl in base.HUE_INFO}
    last_near_hit_ts_by_label["__frame__"] = datetime.min
    last_global_label = None

    print(f"Processing {date_f}/{hr}/{clip_f}")
    cap = cv2.VideoCapture(clip_path)
    if not cap.isOpened():
        print(f"Can't open {clip_path}")
        return [], []

    clip_start_time = time.time()
    frames_in_clip = 0

    w_frame = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h_frame = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    frame_area = w_frame * h_frame

    # FPS (for dynamic skip computation)
    clip_fps = cap.get(cv2.CAP_PROP_FPS)
    try:
        clip_fps = float(clip_fps)
    except Exception:
        clip_fps = 20.0
    if clip_fps <= 0 or np.isnan(clip_fps):
        clip_fps = 20.0

    frame_count = 0

    # Read all frames, but also remember the last frame for post-loop analysis
    last_frame = None
    last_ms = None
    last_vid_secs = None
    last_real_ts = None
    last_frame_skipped = False
    
    while True:
        ret, frame = cap.read()
        if not ret:
            # Video ended - this is our last frame to process
            # (Note: frame is already released, we'll read the last retained frame below)
            break
        frames_in_clip += 1

        # Timestamp
        ms = cap.get(cv2.CAP_PROP_POS_MSEC)
        vid_secs = ms / 1000.0
        real_ts = clip_start + timedelta(seconds=vid_secs)

        # --- Frame analysis logic ---
        # Dynamic frame skipping for non-STOP cards
        frame_count += 1
        process_this_frame = False
        if real_ts <= focus_active_until:
            effective_skip = 1
        else:
            effective_skip = BASELINE_SKIP_STOP if last_global_label == "STOP" else BASELINE_SKIP_OTHER
        if frame_count % effective_skip == 0:
            process_this_frame = True
            last_frame_skipped = False
        else:
            last_frame_skipped = True
            # Save the last skipped frame info for post-processing
            last_frame = frame.copy()
            last_ms = ms
            last_vid_secs = vid_secs
            last_real_ts = real_ts

        if process_this_frame:
            # Resize frame for faster processing
            if base.RESIZE_FACTOR != 1.0:
                frame = cv2.resize(frame, (0, 0), fx=base.RESIZE_FACTOR, fy=base.RESIZE_FACTOR)
                w_frame_resized = int(w_frame * base.RESIZE_FACTOR)
                h_frame_resized = int(h_frame * base.RESIZE_FACTOR)
                frame_area = w_frame_resized * h_frame_resized

            hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

            near_hit_added = False
            if real_ts > focus_active_until:
                frame_glare_intensity = _calculate_glare_ratio(hsv)
                if frame_glare_initialized:
                    frame_glare_delta = frame_glare_intensity - previous_glare_ratio["__frame__"]
                    if frame_glare_delta >= glare_delta_threshold:
                        # Don't create glare entry if there was a recent confirmation
                        if (real_ts - last_confirmation_ts).total_seconds() > GLARE_CONFIRMATION_DEBOUNCE_SECS:
                            last_nh = last_near_hit_ts_by_label.get("__frame__", datetime.min)
                            if (real_ts - last_nh).total_seconds() >= NEAR_HIT_DEBOUNCE_SECS:
                                if not _has_recent_near_hit("__frame__", real_ts):
                                    local_near_hits.append({
                                        "Label": "GLARE",
                                        "Closest Label": "GLARE",
                                        "Video Time": f"{vid_secs:.2f}",
                                        "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                                    })
                                    _record_glare_entry(len(local_near_hits) - 1, real_ts)
                                    recent_near_hits_by_label["__frame__"].append(real_ts)
                                    last_near_hit_ts_by_label["__frame__"] = real_ts
                                    focus_active_until = real_ts + timedelta(seconds=NEAR_HIT_FOCUS_SECS)
                                    near_hit_added = True
                else:
                    frame_glare_initialized = True
                previous_glare_ratio["__frame__"] = frame_glare_intensity

            # --- IMMEDIATE STOP DETECTION (every frame) ---
            stop_mask = cv2.bitwise_or(
                _in_range(hsv, base.COLOR_RANGES["STOP"][0], base.COLOR_RANGES["STOP"][1]),
                _in_range(hsv, base.COLOR_RANGES["STOP"][2], base.COLOR_RANGES["STOP"][3])
            )

            stop_cnts, _ = cv2.findContours(stop_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            stop_valid_quads = []
            stop_areas = {}

            for c in stop_cnts:
                area = cv2.contourArea(c)
                stop_areas[id(c)] = area
                if area < MIN_AREA_RATIO * frame_area:
                    continue
                peri = cv2.arcLength(c, True)
                approx = cv2.approxPolyDP(c, 0.04 * peri, True)
                if len(approx) == 4:
                    x, y, w, h = cv2.boundingRect(c)
                    if area / (w * h) >= 0.70:
                        stop_valid_quads.append(c)

            stop_valid_cnt = max(stop_valid_quads, key=lambda c: stop_areas[id(c)]) if stop_valid_quads else None

            if stop_valid_cnt is not None:
                stop_consec += 1
                if stop_consec >= MIN_FRAMES:
                    if not stop_active:
                        prev = last_detected["STOP"]
                        if (real_ts - prev).total_seconds() > CONFIRMATION_DEBOUNCE_SECS:
                            local_detections.append({
                                "Label": "STOP",
                                "Video Time": f"{vid_secs:.2f}",
                                "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                            })
                            _remove_recent_glare(real_ts)
                            last_detected["STOP"] = real_ts
                            last_global_detected = real_ts
                            last_global_label = "STOP"
                            for k in consec_counts:
                                consec_counts[k] = 0
                            focus_active_until = real_ts + timedelta(seconds=POST_CONFIRM_FOCUS_SECS)
                    stop_active = True
                    continue
            else:
                stop_consec = 0
                stop_active = False

            # --- PROCESS OTHER CARDS in priority order ---
            for lbl in ["MEETING", "START", "DOWN", "LOOK"]:
                hc, hrange = base.HUE_INFO[lbl]
                area_thresh = MIN_AREA_RATIO
                frame_thresh = MIN_FRAMES

                # Build mask for this label
                if lbl == "MEETING":
                    hsv_mask = _in_range(hsv, base.COLOR_RANGES["MEETING"][0], base.COLOR_RANGES["MEETING"][1])
                    raw_mask = cv2.bitwise_and(hsv_mask, cv2.bitwise_not(stop_mask))
                    v = hsv[:, :, 2]
                    _, v_mask = cv2.threshold(v, 80, 255, cv2.THRESH_BINARY)
                    raw_mask = cv2.bitwise_and(raw_mask, v_mask)
                else:
                    lo, hi = base.COLOR_RANGES[lbl]
                    raw_mask = _in_range(hsv, lo, hi)

                mask = cv2.morphologyEx(raw_mask, cv2.MORPH_CLOSE, base.kernel)

                cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                valid_quads = []
                contour_areas = {}

                for c in cnts:
                    area = cv2.contourArea(c)
                    contour_areas[id(c)] = area
                    if area < area_thresh * frame_area:
                        continue
                    peri = cv2.arcLength(c, True)
                    approx = cv2.approxPolyDP(c, 0.04 * peri, True)
                    if len(approx) == 4:
                        valid_quads.append(c)

                valid_cnt = max(valid_quads, key=lambda c: contour_areas[id(c)]) if valid_quads else None

                if lbl == "MEETING" and valid_cnt is not None:
                    area = contour_areas[id(valid_cnt)]
                    if area > 0.6 * frame_area:
                        valid_cnt = None

                # Evaluate glare on the best available contour (even if not a quad)
                glare_candidate_cnt = valid_cnt
                if glare_candidate_cnt is None and cnts:
                    glare_candidate_cnt = max(cnts, key=cv2.contourArea)
                    if cv2.contourArea(glare_candidate_cnt) < area_thresh * frame_area:
                        glare_candidate_cnt = None

                has_glare_spike = False
                if glare_candidate_cnt is not None:
                    xg, yg, wg, hg = cv2.boundingRect(glare_candidate_cnt)
                    glare_roi_hsv = hsv[yg:yg + hg, xg:xg + wg]
                    roi_glare_intensity = _calculate_glare_ratio(glare_roi_hsv) if glare_roi_hsv.size > 0 else 0.0
                    glare_delta = roi_glare_intensity - previous_glare_ratio[lbl]
                    has_glare_spike = glare_delta >= glare_delta_threshold
                    previous_glare_ratio[lbl] = roi_glare_intensity

                consec_counts[lbl] = consec_counts[lbl] + 1 if valid_cnt is not None else 0

                if consec_counts[lbl] < frame_thresh or valid_cnt is None:
                    continue

                # We have a valid candidate
                x, y, w_bb, h_bb = cv2.boundingRect(valid_cnt)
                roi = frame[y:y + h_bb, x:x + w_bb]

                # Optimization: Calculate hue score first to skip expensive template matching for weak colors
                roi_hsv = hsv[y:y + h_bb, x:x + w_bb]
                avg_hue = float(np.mean(roi_hsv[:, :, 0]))
                diff = abs(avg_hue - hc)
                if lbl == "STOP":
                    diff = min(diff, 180 - diff)
                hue_score = max(0.0, 1.0 - diff / hrange)
                
                # If hue is very weak, skip expensive template/edge matching and OCR entirely
                if hue_score < 0.15:
                    consec_counts[lbl] = 0
                    continue

                # Prepare ROI for template matching and OCR
                gray_roi = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
                _, roi_bin = cv2.threshold(gray_roi, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                roi_rs = cv2.resize(roi_bin, (base.template_w, base.template_h), interpolation=cv2.INTER_AREA)

                # Template and edge matching
                roi_edges = cv2.Canny(roi_rs, 50, 150)
                gray_corr = cv2.matchTemplate(roi_rs, base.templates[lbl], cv2.TM_CCOEFF_NORMED).max()
                edge_corr = cv2.matchTemplate(roi_edges, base.edge_templates[lbl], cv2.TM_CCOEFF_NORMED).max()
                
                # Optimization: Skip expensive OCR if templates are very weak (both < 0.20)
                ocr_ratio = 0.0
                ocr_text = ""
                if gray_corr >= 0.20 or edge_corr >= 0.20 or hue_score >= HUE_STRONG:
                    ocr_raw = pytesseract.image_to_string(
                        roi_rs,
                        config="--psm 8 --oem 3 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                    )
                    ocr_text = "".join(ch for ch in ocr_raw if ch.isalpha()).upper()
                    ocr_ratio = _lev_ratio(ocr_text, lbl) if ocr_text else 0.0

                    if ocr_text:
                        bonus = _ocr_signature_bonus(ocr_text, lbl)
                        penalty = _ocr_signature_penalty(ocr_text, lbl)
                        ocr_ratio = min(1.0, max(0.0, ocr_ratio + bonus - penalty))

                    if ocr_text in base.VALID_WORDS and ocr_ratio >= OCR_STRONG:
                        detected_label = lbl
                        prev = last_detected[detected_label]
                        if (real_ts - prev).total_seconds() > CONFIRMATION_DEBOUNCE_SECS:
                            local_detections.append({
                                "Label": detected_label,
                                "Video Time": f"{vid_secs:.2f}",
                                "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                            })
                            _remove_recent_glare(real_ts)
                            last_detected[detected_label] = real_ts
                            last_global_detected = real_ts
                            last_global_label = detected_label
                            for k in consec_counts:
                                consec_counts[k] = 0
                            focus_active_until = real_ts + timedelta(seconds=POST_CONFIRM_FOCUS_SECS)
                        break

                final_score = (
                    base.W_TEMPLATE * gray_corr +
                    base.W_EDGE * edge_corr +
                    base.W_HUE * hue_score +
                    base.W_OCR * ocr_ratio
                )

                # Accept if any scoring method is strong enough
                if (gray_corr >= TEMPLATE_STRONG or
                        edge_corr >= base.EDGE_STRONG or
                        final_score >= base.THRESHOLD or
                        hue_score >= HUE_STRONG):
                    prev = last_detected[lbl]
                    if (real_ts - prev).total_seconds() > CONFIRMATION_DEBOUNCE_SECS:
                        local_detections.append({
                            "Label": lbl,
                            "Video Time": f"{vid_secs:.2f}",
                            "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                        })
                        _remove_recent_glare(real_ts)
                        last_detected[lbl] = real_ts
                        last_global_detected = real_ts
                        last_global_label = lbl
                        for k in consec_counts:
                            consec_counts[k] = 0
                        focus_active_until = real_ts + timedelta(seconds=POST_CONFIRM_FOCUS_SECS)
                    break

                # Near-hit detection
                if valid_cnt is not None:
                    near_hue_ok = hue_score >= LOOSE_HUE_FOCUS
                    sum_no_ocr = (base.W_TEMPLATE + base.W_EDGE + base.W_HUE)
                    final_no_ocr = (
                        (base.W_TEMPLATE * gray_corr) +
                        (base.W_EDGE * edge_corr) +
                        (base.W_HUE * hue_score)
                    ) / sum_no_ocr if sum_no_ocr > 0 else 0.0
                    
                    near_score_band = 0.20 <= final_no_ocr < base.THRESHOLD
                    near_edge_ok = (edge_corr >= 0.30 and edge_corr < base.EDGE_STRONG)
                    near_template_ok = (gray_corr >= 0.20 and gray_corr < TEMPLATE_STRONG)
                    
                    # Glare does not require hue gating
                    if (near_hue_ok and (near_score_band or near_edge_ok or near_template_ok)) or has_glare_spike:
                        last_nh = last_near_hit_ts_by_label.get(lbl, datetime.min)
                        if (real_ts - last_nh).total_seconds() >= NEAR_HIT_DEBOUNCE_SECS:
                            if not _has_recent_near_hit(lbl, real_ts):
                                # Check if there was a recent confirmation before creating glare entry
                                if has_glare_spike and (real_ts - last_confirmation_ts).total_seconds() <= GLARE_CONFIRMATION_DEBOUNCE_SECS:
                                    continue  # Skip this glare entry
                                if not has_glare_spike:
                                    _override_recent_glare(lbl, real_ts)
                                near_label = "GLARE" if has_glare_spike else "NEED CONFIRMATION"
                                # Always log as NEED CONFIRMATION (don't spam GLARE entries)
                                local_near_hits.append({
                                    "Label": near_label,
                                    "Closest Label": lbl,
                                    "Video Time": f"{vid_secs:.2f}",
                                    "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                                })
                                if has_glare_spike:
                                    _record_glare_entry(len(local_near_hits) - 1, real_ts)
                                recent_near_hits_by_label[lbl].append(real_ts)
                                last_near_hit_ts_by_label[lbl] = real_ts
                                focus_active_until = real_ts + timedelta(seconds=NEAR_HIT_FOCUS_SECS)
                                near_hit_added = True

            # --- GENERIC RECTANGLE NEAR-HIT (color-agnostic) ---
            # Only run during focus mode (expensive operation)
            if not near_hit_added and real_ts <= focus_active_until:
                gray_full = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                edges_full = cv2.Canny(gray_full, 50, 150)
                cnts_full, _ = cv2.findContours(edges_full, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                rect_candidates = []
                for c in cnts_full:
                    area = cv2.contourArea(c)
                    if area < MIN_AREA_RATIO * frame_area:
                        continue
                    peri = cv2.arcLength(c, True)
                    approx = cv2.approxPolyDP(c, 0.03 * peri, True)
                    x2, y2, w2, h2 = cv2.boundingRect(c)
                    solidity = area / (w2 * h2) if (w2 * h2) > 0 else 0
                    if len(approx) >= 4 and solidity >= 0.6:
                        rect_candidates.append((c, area))
                if rect_candidates:
                    best_cnt, _ = max(rect_candidates, key=lambda t: t[1])
                    x2, y2, w2, h2 = cv2.boundingRect(best_cnt)
                    guess = base.guess_label_from_bbox(hsv, x2, y2, w2, h2)
                    last_nh = last_near_hit_ts_by_label.get(guess, datetime.min)
                    if (real_ts - last_nh).total_seconds() >= NEAR_HIT_DEBOUNCE_SECS:
                        if not _has_recent_near_hit(guess, real_ts):
                            roi_hsv = hsv[y2:y2 + h2, x2:x2 + w2]
                            roi_glare_intensity = _calculate_glare_ratio(roi_hsv) if roi_hsv.size > 0 else 0.0
                            glare_delta = roi_glare_intensity - previous_glare_ratio.get(guess, 0.0)
                            has_glare_spike = glare_delta >= glare_delta_threshold
                            previous_glare_ratio[guess] = roi_glare_intensity
                            # Check if there was a recent confirmation before creating glare entry
                            if not (has_glare_spike and (real_ts - last_confirmation_ts).total_seconds() <= GLARE_CONFIRMATION_DEBOUNCE_SECS):
                                if not has_glare_spike:
                                    _override_recent_glare(guess, real_ts)
                                near_label = "GLARE" if has_glare_spike else "NEED CONFIRMATION"
                                local_near_hits.append({
                                    "Label": near_label,
                                    "Closest Label": guess,
                                    "Video Time": f"{vid_secs:.2f}",
                                    "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                                })
                                if has_glare_spike:
                                    _record_glare_entry(len(local_near_hits) - 1, real_ts)
                                last_near_hit_ts_by_label[guess] = real_ts



    # After the loop, analyze the last frame if it was skipped due to frame skipping
    # DISABLED for now - skipping post-processing of final frame
    if False and last_frame is not None and last_frame_skipped and last_real_ts is not None:
        # Process the final frame to catch any pending detections
        frame = last_frame
        vid_secs = last_vid_secs
        real_ts = last_real_ts
        
        # Resize frame for faster processing
        frame_resized = frame
        current_frame_area = w_frame * h_frame
        if base.RESIZE_FACTOR != 1.0:
            frame_resized = cv2.resize(frame, (0, 0), fx=base.RESIZE_FACTOR, fy=base.RESIZE_FACTOR)
            w_frame_resized = int(w_frame * base.RESIZE_FACTOR)
            h_frame_resized = int(h_frame * base.RESIZE_FACTOR)
            current_frame_area = w_frame_resized * h_frame_resized
        
        hsv = cv2.cvtColor(frame_resized, cv2.COLOR_BGR2HSV)

        # Check for STOP
        stop_mask = cv2.bitwise_or(
            _in_range(hsv, base.COLOR_RANGES["STOP"][0], base.COLOR_RANGES["STOP"][1]),
            _in_range(hsv, base.COLOR_RANGES["STOP"][2], base.COLOR_RANGES["STOP"][3])
        )
        
        stop_cnts, _ = cv2.findContours(stop_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        stop_valid_quads = []
        stop_areas = {}
        
        for c in stop_cnts:
            area = cv2.contourArea(c)
            stop_areas[id(c)] = area
            if area < MIN_AREA_RATIO * current_frame_area:
                continue
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.04 * peri, True)
            if len(approx) == 4:
                x, y, w, h = cv2.boundingRect(c)
                if area / (w * h) >= 0.70:
                    stop_valid_quads.append(c)
        
        if stop_valid_quads:
            prev = last_detected["STOP"]
            if (real_ts - prev).total_seconds() > CONFIRMATION_DEBOUNCE_SECS:
                local_detections.append({
                    "Label": "STOP",
                    "Video Time": f"{vid_secs:.2f}",
                    "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                })
                last_detected["STOP"] = real_ts
        
        # Check for other cards
        for lbl in ["MEETING", "START", "DOWN", "LOOK"]:
            hc, hrange = base.HUE_INFO[lbl]
            
            # Build mask for this label
            if lbl == "MEETING":
                hsv_mask = _in_range(hsv, base.COLOR_RANGES["MEETING"][0], base.COLOR_RANGES["MEETING"][1])
                raw_mask = cv2.bitwise_and(hsv_mask, cv2.bitwise_not(stop_mask))
                v = hsv[:, :, 2]
                _, v_mask = cv2.threshold(v, 80, 255, cv2.THRESH_BINARY)
                raw_mask = cv2.bitwise_and(raw_mask, v_mask)
            else:
                lo, hi = base.COLOR_RANGES[lbl]
                raw_mask = _in_range(hsv, lo, hi)
            
            mask = cv2.morphologyEx(raw_mask, cv2.MORPH_CLOSE, base.kernel)
            
            cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            valid_quads = []
            contour_areas = {}
            
            for c in cnts:
                area = cv2.contourArea(c)
                contour_areas[id(c)] = area
                if area < MIN_AREA_RATIO * current_frame_area:
                    continue
                peri = cv2.arcLength(c, True)
                approx = cv2.approxPolyDP(c, 0.04 * peri, True)
                if len(approx) == 4:
                    valid_quads.append(c)
            
            if not valid_quads:
                continue
            
            valid_cnt = max(valid_quads, key=lambda c: contour_areas[id(c)])
            
            if lbl == "MEETING":
                area = contour_areas[id(valid_cnt)]
                if area > 0.6 * current_frame_area:
                    continue
            
            # We have a valid candidate - check if we should log it
            x, y, w_bb, h_bb = cv2.boundingRect(valid_cnt)
            roi = frame_resized[y:y + h_bb, x:x + w_bb]
            
            # OCR preparation
            gray_roi = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            _, roi_bin = cv2.threshold(gray_roi, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            roi_rs = cv2.resize(roi_bin, (base.template_w, base.template_h), interpolation=cv2.INTER_AREA)
            
            # OCR
            ocr_raw = pytesseract.image_to_string(
                roi_rs,
                config="--psm 8 --oem 3 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            )
            ocr_text = "".join(ch for ch in ocr_raw if ch.isalpha()).upper()
            ocr_ratio = _lev_ratio(ocr_text, lbl) if ocr_text else 0.0
            
            # Apply signature bonuses
            if ocr_text:
                bonus = _ocr_signature_bonus(ocr_text, lbl)
                penalty = _ocr_signature_penalty(ocr_text, lbl)
                ocr_ratio = min(1.0, max(0.0, ocr_ratio + bonus - penalty))
            
            # Strong OCR detection
            if ocr_text in base.VALID_WORDS and ocr_ratio >= OCR_STRONG:
                prev = last_detected[lbl]
                if (real_ts - prev).total_seconds() > CONFIRMATION_DEBOUNCE_SECS:
                    local_detections.append({
                        "Label": lbl,
                        "Video Time": f"{vid_secs:.2f}",
                        "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                    })
                    last_detected[lbl] = real_ts
                continue
            
            # Fallback scoring
            roi_edges = cv2.Canny(roi_rs, 50, 150)
            gray_corr = cv2.matchTemplate(roi_rs, base.templates[lbl], cv2.TM_CCOEFF_NORMED).max()
            edge_corr = cv2.matchTemplate(roi_edges, base.edge_templates[lbl], cv2.TM_CCOEFF_NORMED).max()
            avg_hue = float(np.mean(hsv[y:y + h_bb, x:x + w_bb, 0]))
            diff = abs(avg_hue - hc)
            if lbl == "STOP":
                diff = min(diff, 180 - diff)
            hue_score = max(0.0, 1.0 - diff / hrange)
            
            final_score = (
                base.W_TEMPLATE * gray_corr +
                base.W_EDGE * edge_corr +
                base.W_HUE * hue_score +
                base.W_OCR * ocr_ratio
            )
            
            # Accept if any scoring method is strong enough
            if (gray_corr >= TEMPLATE_STRONG or
                    edge_corr >= base.EDGE_STRONG or
                    final_score >= base.THRESHOLD or
                    hue_score >= HUE_STRONG):
                prev = last_detected[lbl]
                if (real_ts - prev).total_seconds() > CONFIRMATION_DEBOUNCE_SECS:
                    local_detections.append({
                        "Label": lbl,
                        "Video Time": f"{vid_secs:.2f}",
                        "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                    })
                    last_detected[lbl] = real_ts

        # Generic rectangle near-hit for last frame
        gray_full = cv2.cvtColor(frame_resized, cv2.COLOR_BGR2GRAY)
        edges_full = cv2.Canny(gray_full, 50, 150)
        cnts_full, _ = cv2.findContours(edges_full, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        rect_candidates = []
        for c in cnts_full:
            area = cv2.contourArea(c)
            if area < MIN_AREA_RATIO * current_frame_area:
                continue
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.03 * peri, True)
            x2, y2, w2, h2 = cv2.boundingRect(c)
            solidity = area / (w2 * h2) if (w2 * h2) > 0 else 0
            if len(approx) >= 4 and solidity >= 0.6:
                rect_candidates.append((c, area))
        if rect_candidates:
            best_cnt, _ = max(rect_candidates, key=lambda t: t[1])
            x2, y2, w2, h2 = cv2.boundingRect(best_cnt)
            guess = base.guess_label_from_bbox(hsv, x2, y2, w2, h2)
            last_nh = last_near_hit_ts_by_label.get(guess, datetime.min)
            if (real_ts - last_nh).total_seconds() >= NEAR_HIT_DEBOUNCE_SECS:
                if not _has_recent_near_hit(guess, real_ts):
                    roi_hsv = hsv[y2:y2 + h2, x2:x2 + w2]
                    near_label = "GLARE" if _is_glare(roi_hsv) else "NEED CONFIRMATION"
                    local_near_hits.append({
                        "Label": near_label,
                        "Closest Label": guess,
                        "Video Time": f"{vid_secs:.2f}",
                        "Real Time": real_ts.strftime("%Y-%m-%d %H:%M:%S")
                    })
                    last_near_hit_ts_by_label[guess] = real_ts



    cap.release()

    # Stats (optional)
    clip_elapsed = time.time() - clip_start_time
    fps = frames_in_clip / clip_elapsed if clip_elapsed > 0 else 0
    if base.VERBOSE:
        print(f"    Processed {frames_in_clip} frames in {clip_elapsed:.2f}s ({fps:.1f} fps)")

    combined = local_detections + local_near_hits
    return combined


def auto_adjust_xlsx_columns(writer, sheet_name, df):
    """
    Auto-adjust columns width in an Excel worksheet to fit the content.
    Also:
    - Removes all cell borders
    - Centers the table on the page for printing
    - Keeps headers bold
    - Applies proper formatting to date and time columns
    """
    worksheet = writer.sheets[sheet_name]
    
    # Auto-adjust columns based on DataFrame content width
    for idx, col in enumerate(df.columns):
        # Find the maximum length in the column
        max_len = max(
            # Column name length
            len(str(col)),
            # Maximum content length
            df[col].astype(str).str.len().max() if not df.empty else 0
        )
        
        # Add some padding
        adjusted_width = max_len + 2
        
        # Limit width to a reasonable maximum
        if adjusted_width > 50:
            adjusted_width = 50
        
        # Set the column width
        col_letter = chr(65 + idx) if idx < 26 else chr(64 + idx//26) + chr(65 + idx%26)
        worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Apply proper formatting to Time column
        if col == "Time":
            for row_idx, _ in enumerate(df.index, start=2):
                cell = f"{col_letter}{row_idx}"
                worksheet[cell].number_format = "hh:mm:ss"
                
        # Apply proper formatting to Date column
        elif col == "Date":
            for row_idx, _ in enumerate(df.index, start=2):
                cell = f"{col_letter}{row_idx}"
                worksheet[cell].number_format = "mm/dd/yyyy"

    # Remove borders from all cells
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    
    # Apply formatting to all cells
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = no_border
            
            # Keep headers bold
            if row == 1:
                cell.font = Font(bold=True)
    
    # Center the table for printing
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = True
    worksheet.page_setup.fitToHeight = True
    worksheet.print_options.horizontalCentered = True
    worksheet.print_options.verticalCentered = True


def add_time_borders(worksheet, df, columns):
    """
    Add borders to separate detections by hour and day.
    - Regular bottom border when hour changes
    - Thick bottom border when day changes
    Preserves existing cell formatting (fills, fonts, etc.)
    """
    # Skip if empty dataframe
    if df.empty:
        return
        
    # Create border styles with all sides preserved
    thin_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thick')
    )
    
    # Sort by time to ensure chronological order
    df = df.sort_values("Real Time")
    
    # Get the total number of columns with data
    max_col = len(columns)
    
    # Start from row 2 (first data row after header)
    prev_hour = df.iloc[0]["Real Time"].hour
    prev_day = df.iloc[0]["Real Time"].day
    
    # Iterate through rows (starting from index 1, row 3 in Excel)
    for i in range(1, len(df)):
        current_hour = df.iloc[i]["Real Time"].hour
        current_day = df.iloc[i]["Real Time"].day
        excel_row = i + 2  # +2 because Excel is 1-indexed and we have a header row
        
        # Check for day change (stronger visual separation)
        if current_day != prev_day:
            # Add thick border to previous row (preserving other properties)
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=excel_row-1, column=col)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=Side(style='thick')
                )
            prev_hour = current_hour
            prev_day = current_day
            
        # Check for hour change
        elif current_hour != prev_hour:
            # Add thin border to previous row (preserving other properties)
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=excel_row-1, column=col)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=Side(style='thin')
                )
            prev_hour = current_hour


def _apply_near_hit_formatting(worksheet, start_row, end_row, max_col):
    """
    Apply Excel conditional formatting to highlight GLARE and NEED CONFIRMATION rows.
    This allows the highlighting to automatically update when users edit labels.
    """
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    dxf = DifferentialStyle(fill=yellow_fill)
    
    # Apply to all columns (A through the last column)
    last_col_letter = get_column_letter(max_col)
    
    # Rule 1: Highlight when column A contains "GLARE"
    rule_glare = Rule(type="expression", dxf=dxf)
    rule_glare.formula = [f'$A{start_row}="GLARE"']
    worksheet.conditional_formatting.add(f'A{start_row}:{last_col_letter}{end_row}', rule_glare)
    
    # Rule 2: Highlight when column A contains "NEED CONFIRMATION"
    rule_need_confirm = Rule(type="expression", dxf=dxf)
    rule_need_confirm.formula = [f'$A{start_row}="NEED CONFIRMATION"']
    worksheet.conditional_formatting.add(f'A{start_row}:{last_col_letter}{end_row}', rule_need_confirm)


def main():
    """
    Process all video clips using the fast analyzer and generate Excel report
    using fast analyzer settings (not base analyzer settings).
    """
    root_folder = base.root_folder
    output_excel = base.output_excel
    
    # Start timer for overall processing
    start_time = time.time()
    
    print(f"🎯 Time Study Analyzer (Fast) {base.__version__}")
    print("=" * 40)
    
    # Collect all clips to process
    clips_to_process = []
    
    for date_f in sorted(os.listdir(root_folder)):
        date_path = os.path.join(root_folder, date_f)
        if not os.path.isdir(date_path) or not date_f.isdigit():
            continue

        for hr in sorted(os.listdir(date_path)):
            hr_path = os.path.join(date_path, hr)
            if not os.path.isdir(hr_path) or not hr.isdigit():
                continue

            for clip_f in sorted(os.listdir(hr_path)):
                if not clip_f.endswith(".mp4") or not clip_f[:-4].isdigit():
                    continue
                    
                clip_date = datetime.strptime(date_f, "%Y%m%d")
                clip_start = clip_date + timedelta(
                    hours=int(hr), minutes=int(clip_f[:-4])
                )
                clip_path = os.path.join(hr_path, clip_f)
                
                clips_to_process.append((date_f, hr, clip_f, clip_path, clip_start))
    
    # Process clips in parallel
    all_detections = []
    total_videos_processed = len(clips_to_process)
    
    # Use 60% of CPUs (conservative)
    total_cpus = multiprocessing.cpu_count()
    max_workers_conservative = max(1, int(total_cpus * 0.60))
    num_workers = min(max_workers_conservative, len(clips_to_process))
    print(f"\n🖥️ Processing {total_videos_processed} videos using {num_workers}/{total_cpus} CPU cores")
    
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        results = list(executor.map(process_clip, clips_to_process))
        
    # Combine all results
    for clip_detections in results:
        all_detections.extend(clip_detections)
    
    # Complete processing time
    total_elapsed = time.time() - start_time
    
    print("\n==== Processing Complete ====")
    print(f"Total time: {total_elapsed:.2f} seconds")
    print(f"Videos processed: {total_videos_processed}")
    print(f"Detections found: {len(all_detections)}")
    
    # ─── EXPORT TO EXCEL ─────────────────────────────────────────────────────────
    if all_detections:
        # Process raw data
        df = pd.DataFrame(all_detections)
        real_time = pd.to_datetime(df["Real Time"])
        df["Real Time"] = real_time
        
        # Format dates and times
        df["Date"] = real_time.dt.strftime("%m/%d/%Y")  # M/D/YYYY format
        df["Time"] = real_time.dt.strftime("%H:%M:%S")  # 24-hour format
        
        # Sort by time to ensure chronological order
        df = df.sort_values("Real Time")

        # Create summary statistics dataframe (will be replaced with formulas)
        stats = []
        stats.append({"Metric": "START Time", "Value": ""})
        stats.append({"Metric": "DOWN Time", "Value": ""})
        stats.append({"Metric": "MEETING Time", "Value": ""})
        stats_df = pd.DataFrame(stats)
        
        # Extract LOOK events
        look_events = df[df["Label"] == "LOOK"].copy()
        
        # Prepare Excel writer
        os.makedirs(os.path.dirname(output_excel), exist_ok=True)
        if os.path.exists(output_excel):
            os.remove(output_excel)
            
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            # Tab 1: Summary statistics (using formulas that reference Raw Detections)
            stats_df.to_excel(writer, sheet_name="Summary Statistics", index=False)
            
            # Get the worksheet and replace values with formulas
            summary_ws = writer.sheets["Summary Statistics"]
            
            auto_adjust_xlsx_columns(writer, "Summary Statistics", stats_df)
            
            # Tab 2: LOOK events
            if not look_events.empty:
                # Add video file links for LOOK events
                look_events["Video File"] = look_events.apply(
                    lambda row: f'=HYPERLINK("D:\\record\\{row["Real Time"].strftime("%Y%m%d")}\\{row["Real Time"].strftime("%H")}\\{row["Real Time"].strftime("%M")}.mp4", "Open Video")',
                    axis=1
                )
                
                # Add the Video File column to the output
                look_events.to_excel(writer, sheet_name="LOOK Events", 
                                   index=False, columns=["Date", "Time", "Video Time", "Video File"])
                auto_adjust_xlsx_columns(writer, "LOOK Events", 
                                   look_events[["Date", "Time", "Video Time", "Video File"]])
            else:
                # Make sure empty dataframe has all columns
                empty_df = pd.DataFrame(columns=["Date", "Time", "Video Time", "Video File"])
                empty_df.to_excel(writer, sheet_name="LOOK Events", index=False)
                auto_adjust_xlsx_columns(writer, "LOOK Events", empty_df)
            
            # Tab 3: Raw data (all detections with split date/time)
            # Create video file links
            df["Video File"] = df.apply(
                lambda row: f'=HYPERLINK("D:\\record\\{row["Real Time"].strftime("%Y%m%d")}\\{row["Real Time"].strftime("%H")}\\{row["Real Time"].strftime("%M")}.mp4", "Open Video")',
                axis=1
            )

            # Separate confirmed detections from near-hits
            confirmed = df[~df["Label"].isin(["GLARE", "NEED CONFIRMATION"])].copy()
            near_hits = df[df["Label"].isin(["GLARE", "NEED CONFIRMATION"])].copy()
            
            # If we have both confirmed and near-hits, combine them for the raw output
            if not confirmed.empty or not near_hits.empty:
                combined_data = pd.concat([confirmed, near_hits], ignore_index=True).sort_values("Real Time")
            else:
                # If both are empty, create minimal dataframe with required columns
                combined_data = df.copy()
            
            # Add placeholder column for Duration (will be replaced with formulas)
            combined_data["Duration (sec)"] = 0
            
            # Determine columns to display (include Real Time for duration formulas)
            if "Closest Label" in combined_data.columns:
                # We have near-hit data with Closest Label
                columns = ["Label", "Closest Label", "Date", "Time", "Real Time", "Video Time", "Best Guess", "Video File", "Duration (sec)"]
            else:
                columns = ["Label", "Date", "Time", "Real Time", "Video Time", "Video File", "Duration (sec)"]
            
            # Filter to only existing columns (Video File will be added by pandas, Duration (sec) was just added)
            columns = [col for col in columns if col in combined_data.columns or col in ["Video File", "Duration (sec)"]]

            duration_col_letter = get_column_letter(columns.index("Duration (sec)") + 1)
            real_time_col_letter = get_column_letter(columns.index("Real Time") + 1)
            duration_range = f"'Raw Detections'!{duration_col_letter}:{duration_col_letter}"
            label_range = "'Raw Detections'!A:A"

            # Detected time totals (automatic, read-only)
            summary_ws['B2'] = f"=SUMIFS({duration_range},{label_range},\"START\")/86400"
            summary_ws['B3'] = f"=SUMIFS({duration_range},{label_range},\"DOWN\")/86400"
            summary_ws['B4'] = f"=SUMIFS({duration_range},{label_range},\"MEETING\")/86400"

            for cell in ["B2", "B3", "B4"]:
                summary_ws[cell].number_format = "[h]:mm:ss"
            
            combined_data.to_excel(writer, sheet_name="Raw Detections", index=False, columns=columns)
            auto_adjust_xlsx_columns(writer, "Raw Detections", combined_data[columns])
            
            # Add Duration formulas that calculate time gap to next row
            raw_ws = writer.sheets["Raw Detections"]
            duration_col_idx = columns.index("Duration (sec)") + 1
            duration_col = get_column_letter(duration_col_idx)
            real_time_col_idx = columns.index("Real Time") + 1
            real_time_col = get_column_letter(real_time_col_idx)
            
            for row_num in range(2, len(combined_data) + 2):
                next_row = row_num + 1
                # Calculate duration as time difference to next row (in seconds)
                # Formula: If label is START/DOWN/MEETING, calculate (next_time - current_time) * 86400
                raw_ws[f"{duration_col}{row_num}"] = (
                    f"=IF(OR(A{row_num}=\"START\",A{row_num}=\"DOWN\",A{row_num}=\"MEETING\"),"
                    f"IF(ROW()>={len(combined_data) + 1},0,"
                    f"({real_time_col}{next_row}-{real_time_col}{row_num})*86400),0)"
                )
            
            # Hide Real Time and Duration columns (they're needed for formulas but not for user viewing)
            raw_ws.column_dimensions[real_time_col].hidden = True
            raw_ws.column_dimensions[duration_col].hidden = True
            
            # Apply time borders first
            add_time_borders(writer.sheets["Raw Detections"], combined_data, columns)
            
            # Apply Excel conditional formatting for near-hits (auto-updates when user edits labels)
            near_hit_start_row = 2  # Start after header
            _apply_near_hit_formatting(raw_ws, near_hit_start_row, 
                                      len(combined_data) + 1, len(columns))

        print(f"\n✅ Fast analyzer report saved to {output_excel}")
    else:
        print("\n⚠️ No detections found.")
        
    # OPTIMIZATION: Release memory after processing
    cv2.destroyAllWindows()


if __name__ == "__main__":
    main()
